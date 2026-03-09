"""Standalone PyQt dashboard for DecisionPipeline reflection proposals."""

from __future__ import annotations

import csv
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Sequence

from importlib.util import find_spec

if find_spec("PyQt6") is not None:  # pragma: no cover - runtime dependent
    from PyQt6.QtGui import QColor
    from PyQt6.QtWidgets import (
        QApplication,
        QAbstractItemView,
        QCheckBox,
        QFileDialog,
        QGridLayout,
        QGroupBox,
        QHBoxLayout,
        QLabel,
        QListWidget,
        QListWidgetItem,
        QMessageBox,
        QPushButton,
        QTableWidget,
        QTableWidgetItem,
        QTextEdit,
        QVBoxLayout,
        QWidget,
    )
else:  # pragma: no cover - runtime dependent
    from PyQt5.QtGui import QColor
    from PyQt5.QtWidgets import (
        QApplication,
        QAbstractItemView,
        QCheckBox,
        QFileDialog,
        QGridLayout,
        QGroupBox,
        QHBoxLayout,
        QLabel,
        QListWidget,
        QListWidgetItem,
        QMessageBox,
        QPushButton,
        QTableWidget,
        QTableWidgetItem,
        QTextEdit,
        QVBoxLayout,
        QWidget,
    )

try:  # pragma: no cover - runtime import convenience
    from .decision_pipeline import DecisionPipeline
    from .reflection_actions import ReflectionActionHandler
    from .reflection_formatting import format_proposals
except ImportError:  # pragma: no cover - direct script execution
    from agent.decision_pipeline import DecisionPipeline
    from agent.reflection_actions import ReflectionActionHandler
    from agent.reflection_formatting import format_proposals


class NumericTableWidgetItem(QTableWidgetItem):
    """QTableWidgetItem variant with numeric sorting for heatmap columns."""

    def __init__(self, value: float, text: str):
        super().__init__(text)
        self._value = value

    def __lt__(self, other: object) -> bool:
        if isinstance(other, NumericTableWidgetItem):
            return self._value < other._value
        return super().__lt__(other)


class ReflectionWidget(QWidget):
    """Interactive dashboard for reviewing and applying reflection proposals."""

    def __init__(self, pipeline: DecisionPipeline):
        super().__init__()
        self.pipeline = pipeline
        self.action_handler = ReflectionActionHandler(pipeline)
        self.current_proposals: List[Dict[str, Any]] = []
        self.strategy_filters: Dict[str, QCheckBox] = {}
        self.activity_filters: Dict[str, QCheckBox] = {}

        self.setWindowTitle("Reflection Dashboard")
        self.resize(1180, 860)
        self._build_ui()
        self._refresh_dashboard()

    def _build_ui(self) -> None:
        """Create and connect dashboard widgets."""
        root = QVBoxLayout(self)

        control_row = QHBoxLayout()
        self.reflect_button = QPushButton("Reflect & Propose")
        self.reflect_button.clicked.connect(self.on_reflect_clicked)
        control_row.addWidget(self.reflect_button)

        self.approve_button = QPushButton("Approve Selected")
        self.approve_button.clicked.connect(self.on_approve_selected)
        control_row.addWidget(self.approve_button)

        self.reject_button = QPushButton("Reject Selected")
        self.reject_button.clicked.connect(self.on_reject_selected)
        control_row.addWidget(self.reject_button)

        self.export_button = QPushButton("Export Report")
        self.export_button.clicked.connect(self.on_export_report)
        control_row.addWidget(self.export_button)

        root.addLayout(control_row)

        grid = QGridLayout()

        proposal_box = QGroupBox("Proposals")
        proposal_layout = QVBoxLayout(proposal_box)
        self.proposal_summary = QTextEdit()
        self.proposal_summary.setReadOnly(True)
        proposal_layout.addWidget(self.proposal_summary)

        self.proposal_list = QListWidget()
        self.proposal_list.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        proposal_layout.addWidget(self.proposal_list)
        grid.addWidget(proposal_box, 0, 0)

        settings_box = QGroupBox("Current Pipeline Settings")
        settings_layout = QVBoxLayout(settings_box)
        self.threshold_label = QLabel()
        self.fallback_label = QLabel()
        self.sleep_phase_label = QLabel()
        settings_layout.addWidget(self.threshold_label)
        settings_layout.addWidget(self.fallback_label)
        settings_layout.addWidget(self.sleep_phase_label)
        grid.addWidget(settings_box, 0, 1)

        heatmap_box = QGroupBox("Strategy Heatmap")
        heatmap_layout = QVBoxLayout(heatmap_box)

        strategy_filter_row = QHBoxLayout()
        strategy_filter_row.addWidget(QLabel("Filter strategies:"))
        for strategy in ("structured_reasoning", "answer_with_tool", "retrieve_context"):
            checkbox = QCheckBox(strategy)
            checkbox.setChecked(True)
            checkbox.stateChanged.connect(self._refresh_heatmap)
            self.strategy_filters[strategy] = checkbox
            strategy_filter_row.addWidget(checkbox)
        heatmap_layout.addLayout(strategy_filter_row)

        self.heatmap_table = QTableWidget(0, 3)
        self.heatmap_table.setHorizontalHeaderLabels(["Strategy", "Success Rate", "Efficiency"])
        self.heatmap_table.setSortingEnabled(True)
        self.heatmap_table.horizontalHeaderItem(1).setToolTip("Success Rate = successes / attempts")
        self.heatmap_table.horizontalHeaderItem(2).setToolTip("Efficiency = total_value / attempts")
        heatmap_layout.addWidget(self.heatmap_table)
        grid.addWidget(heatmap_box, 1, 0)

        trends_box = QGroupBox("Recent Trends")
        trends_layout = QVBoxLayout(trends_box)
        self.trends_text = QTextEdit()
        self.trends_text.setReadOnly(True)
        trends_layout.addWidget(self.trends_text)
        grid.addWidget(trends_box, 1, 1)

        activity_box = QGroupBox("Pipeline Activity (last 20)")
        activity_layout = QVBoxLayout(activity_box)
        activity_filter_row = QHBoxLayout()
        activity_filter_row.addWidget(QLabel("Action type:"))
        for action_type in ("strategy_mutation", "control_update", "tool_demotion"):
            checkbox = QCheckBox(action_type)
            checkbox.setChecked(True)
            checkbox.stateChanged.connect(self._refresh_activity_log)
            self.activity_filters[action_type] = checkbox
            activity_filter_row.addWidget(checkbox)
        activity_layout.addLayout(activity_filter_row)
        self.activity_text = QTextEdit()
        self.activity_text.setReadOnly(True)
        activity_layout.addWidget(self.activity_text)
        grid.addWidget(activity_box, 2, 0, 1, 2)

        root.addLayout(grid)

        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        root.addWidget(self.log_output)

    def on_reflect_clicked(self) -> None:
        """Request reflection proposals from pipeline and refresh the panel."""
        self.current_proposals = self.pipeline.reflect_and_propose()
        auto_messages = self.action_handler.auto_apply_high_confidence(self.current_proposals)
        self.current_proposals = [item for item in self.current_proposals if float(item.get("confidence", 0.0)) < 0.9]
        self._render_proposals()
        if auto_messages:
            self._append_log("; ".join(auto_messages))
        self._append_log(f"generated {len(self.current_proposals)} manual proposals")
        self._refresh_dashboard()
        self._auto_save_state()

    def on_approve_selected(self) -> None:
        """Apply selected proposals and update metrics/controls immediately."""
        selected = self._selected_proposals()
        if not selected:
            QMessageBox.warning(self, "No Selection", "Please select at least one proposal to approve.")
            return

        messages = self.action_handler.apply_selected(selected)
        self._remove_selected_proposals(selected)
        self._refresh_dashboard()
        self._append_log("; ".join(messages))
        self._auto_save_state()

    def on_reject_selected(self) -> None:
        """Reject selected proposals and clear them from current proposal list."""
        selected = self._selected_proposals()
        if not selected:
            QMessageBox.warning(self, "No Selection", "Please select at least one proposal to reject.")
            return

        messages = self.action_handler.reject_selected(selected)
        self._remove_selected_proposals(selected)
        self._refresh_dashboard()
        self._append_log("; ".join(messages))
        self._auto_save_state()

    def on_export_report(self) -> None:
        """Export dashboard state report to .xlsx when available or CSV fallback."""
        target_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Reflection Report",
            f"reflection_report_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Workbook (*.xlsx);;CSV (*.csv)",
        )
        if not target_path:
            return

        if self.export_report_to(target_path):
            self._append_log(f"report exported: {target_path}")

    def export_report_to(self, target: str | Path) -> bool:
        """Public helper to export report to explicit file path for scripted demos/tests."""
        return self._export_report(Path(target))

    def _selected_proposals(self) -> List[Dict[str, Any]]:
        selected_indices = [self.proposal_list.row(item) for item in self.proposal_list.selectedItems()]
        return [self.current_proposals[index] for index in selected_indices if 0 <= index < len(self.current_proposals)]

    def _remove_selected_proposals(self, proposals: Sequence[Dict[str, Any]]) -> None:
        """Drop handled proposals from in-memory state and re-render list."""
        selected_ids = {item.get("proposal_id") for item in proposals}
        self.current_proposals = [item for item in self.current_proposals if item.get("proposal_id") not in selected_ids]
        self._render_proposals()

    def _render_proposals(self) -> None:
        self.proposal_summary.setPlainText(format_proposals(self.current_proposals))
        self.proposal_list.clear()

        if not self.current_proposals:
            return

        for proposal in self.current_proposals:
            confidence = float(proposal.get("confidence", 0.0))
            text = f"{proposal.get('proposal_id')} | {proposal.get('change_type')} | conf={confidence:.2f}"
            item = QListWidgetItem(text)
            if confidence >= 0.9:
                item.setBackground(QColor("#c8f7c5"))
            elif confidence >= 0.7:
                item.setBackground(QColor("#fff3b0"))
            else:
                item.setBackground(QColor("#f8c8c8"))
            self.proposal_list.addItem(item)

    def _refresh_dashboard(self) -> None:
        """Refresh settings, heatmap, trend data and activity stream after each action."""
        self.threshold_label.setText(f"low_confidence_threshold: {self.pipeline.low_confidence_threshold:.2f}")
        self.fallback_label.setText(f"fallback_action: {self.pipeline.fallback_action}")
        self._refresh_sleep_phase_label()
        self._refresh_heatmap()
        self._refresh_trends()
        self._refresh_activity_log()

    def _refresh_sleep_phase_label(self) -> None:
        """Show last sleep-phase timestamp and action count since that event."""
        history = self.pipeline.cognitive_state.get("action_history", [])
        last_sleep_time = self.pipeline.cognitive_state.get("last_sleep_phase_at")
        if not last_sleep_time:
            for item in reversed(history):
                if item.get("action") == "sleep_phase":
                    last_sleep_time = item.get("timestamp")
                    break

        if not last_sleep_time:
            self.sleep_phase_label.setText("Sleep phase: never")
            return

        actions_since_sleep = 0
        for item in reversed(history):
            if item.get("timestamp") == last_sleep_time:
                break
            actions_since_sleep += 1

        self.sleep_phase_label.setText(f"Sleep phase: {last_sleep_time} | actions since sleep: {actions_since_sleep}")

    def _refresh_heatmap(self) -> None:
        """Recompute filtered/sortable heatmap with success color coding."""
        stats = self.pipeline.cognitive_state.get("strategy_stats", {})
        enabled_filters = {name for name, checkbox in self.strategy_filters.items() if checkbox.isChecked()}
        rows = [item for item in stats.items() if isinstance(item[1], dict) and item[0] in enabled_filters]

        self.heatmap_table.setSortingEnabled(False)
        self.heatmap_table.setRowCount(len(rows))
        for row_idx, (strategy, metrics) in enumerate(rows):
            attempts = float(metrics.get("attempts", 0) or 0)
            successes = float(metrics.get("successes", 0) or 0)
            total_value = float(metrics.get("total_value", 0.0) or 0.0)
            success_rate = successes / attempts if attempts else 0.0
            efficiency = total_value / attempts if attempts else 0.0

            strategy_item = QTableWidgetItem(strategy)
            success_item = NumericTableWidgetItem(success_rate, f"{success_rate:.2f}")
            efficiency_item = NumericTableWidgetItem(efficiency, f"{efficiency:.2f}")

            color = QColor("#c8f7c5") if success_rate > 0.9 else QColor("#fff3b0") if success_rate >= 0.7 else QColor("#f8c8c8")
            for item in (strategy_item, success_item, efficiency_item):
                item.setBackground(color)

            self.heatmap_table.setItem(row_idx, 0, strategy_item)
            self.heatmap_table.setItem(row_idx, 1, success_item)
            self.heatmap_table.setItem(row_idx, 2, efficiency_item)

        self.heatmap_table.setSortingEnabled(True)

    def _refresh_trends(self, window: int = 20) -> None:
        history = self.pipeline.cognitive_state.get("action_history", [])[-window:]
        if not history:
            self.trends_text.setPlainText("No action history yet.")
            return

        marks = "".join("✓" if item.get("success") else "✗" for item in history)
        success_count = sum(1 for item in history if item.get("success") is True)
        rate = success_count / len(history)
        strategy_changes = [a for a in self.pipeline.cognitive_state.get("pipeline_activity", []) if a.get("type") == "strategy_mutation"]
        self.trends_text.setPlainText(
            f"Last {len(history)} actions success trend:\n{marks}\n"
            f"Success rate: {rate:.2f}\n"
            f"Pending proposals: {len(self.current_proposals)}\n"
            f"Strategy changes logged: {len(strategy_changes)}"
        )

    def _refresh_activity_log(self) -> None:
        """Render the latest 20 pipeline activities filtered by selected action types."""
        selected_types = {name for name, checkbox in self.activity_filters.items() if checkbox.isChecked()}
        activity_log = self.pipeline.cognitive_state.get("pipeline_activity", [])
        filtered = [entry for entry in activity_log if entry.get("type") in selected_types]
        latest = filtered[-20:]

        if not latest:
            self.activity_text.setPlainText("No matching activities.")
            return

        lines = [
            f"{entry.get('timestamp', '-') } | {entry.get('type', '-') } | {entry.get('details', {})}"
            for entry in latest
        ]
        self.activity_text.setPlainText("\n".join(lines))

    def _append_log(self, message: str) -> None:
        current = self.log_output.toPlainText().strip()
        new_text = f"{current}\n{message}" if current else message
        self.log_output.setPlainText(new_text)

    def _auto_save_state(self) -> None:
        """Persist cognitive state after dashboard changes for reproducible demos."""
        target = Path("reflection_state.json")
        self.pipeline.save_state(str(target))

    def _export_report(self, target: Path) -> bool:
        """Write report to xlsx when openpyxl exists, else export CSV companion files."""
        target = Path(target)
        try:
            import openpyxl  # type: ignore
        except ImportError:
            self._export_csv_report(target.with_suffix(""))
            QMessageBox.information(self, "Report Export", f"openpyxl not installed, exported CSV bundle at {target.with_suffix('')}")
            return True

        workbook = openpyxl.Workbook()
        settings_sheet = workbook.active
        settings_sheet.title = "settings"
        settings_sheet.append(["low_confidence_threshold", self.pipeline.low_confidence_threshold])
        settings_sheet.append(["fallback_action", self.pipeline.fallback_action])

        heatmap_sheet = workbook.create_sheet("heatmap")
        heatmap_sheet.append(["strategy", "success_rate", "efficiency"])
        for row in self._collect_heatmap_rows():
            heatmap_sheet.append(row)

        proposals_sheet = workbook.create_sheet("proposals")
        proposals_sheet.append(["proposal_id", "change_type", "target", "proposed_value", "confidence", "status"])
        for proposal in self.current_proposals:
            proposals_sheet.append([
                proposal.get("proposal_id"),
                proposal.get("change_type"),
                proposal.get("target"),
                proposal.get("proposed_value"),
                proposal.get("confidence"),
                "pending",
            ])

        activity_sheet = workbook.create_sheet("activity")
        activity_sheet.append(["timestamp", "type", "details"])
        for entry in self.pipeline.cognitive_state.get("pipeline_activity", [])[-20:]:
            activity_sheet.append([entry.get("timestamp"), entry.get("type"), str(entry.get("details", {}))])

        workbook.save(target)
        return True

    def _export_csv_report(self, target_base: Path) -> None:
        """Fallback CSV export for environments without Excel dependency."""
        target_base.parent.mkdir(parents=True, exist_ok=True)

        with (target_base.with_name(f"{target_base.name}_settings.csv")).open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["setting", "value"])
            writer.writerow(["low_confidence_threshold", self.pipeline.low_confidence_threshold])
            writer.writerow(["fallback_action", self.pipeline.fallback_action])

        with (target_base.with_name(f"{target_base.name}_heatmap.csv")).open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["strategy", "success_rate", "efficiency"])
            writer.writerows(self._collect_heatmap_rows())

        with (target_base.with_name(f"{target_base.name}_proposals.csv")).open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["proposal_id", "change_type", "target", "proposed_value", "confidence", "status"])
            for proposal in self.current_proposals:
                writer.writerow([
                    proposal.get("proposal_id"),
                    proposal.get("change_type"),
                    proposal.get("target"),
                    proposal.get("proposed_value"),
                    proposal.get("confidence"),
                    "pending",
                ])

        with (target_base.with_name(f"{target_base.name}_activity.csv")).open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["timestamp", "type", "details"])
            for entry in self.pipeline.cognitive_state.get("pipeline_activity", [])[-20:]:
                writer.writerow([entry.get("timestamp"), entry.get("type"), entry.get("details", {})])

    def _collect_heatmap_rows(self) -> List[List[Any]]:
        """Collect heatmap rows from strategy stats for report export."""
        rows: List[List[Any]] = []
        stats = self.pipeline.cognitive_state.get("strategy_stats", {})
        for strategy, metrics in stats.items():
            if not isinstance(metrics, dict):
                continue
            attempts = float(metrics.get("attempts", 0) or 0)
            successes = float(metrics.get("successes", 0) or 0)
            total_value = float(metrics.get("total_value", 0.0) or 0.0)
            success_rate = successes / attempts if attempts else 0.0
            efficiency = total_value / attempts if attempts else 0.0
            rows.append([strategy, round(success_rate, 4), round(efficiency, 4)])
        return rows


_reflection_dashboard_factory: DashboardFactory = ReflectionWidget


def set_reflection_dashboard_factory(factory: DashboardFactory) -> None:
    """Set injectable dashboard factory for GhostGPT integration and testing."""
    global _reflection_dashboard_factory
    _reflection_dashboard_factory = factory


def create_reflection_dashboard(pipeline: DecisionPipeline) -> QWidget:
    """Build reflection dashboard from configurable factory."""
    return _reflection_dashboard_factory(pipeline)


def build_demo_pipeline() -> DecisionPipeline:
    """Build demo pipeline with synthetic cognitive state for quick local testing."""
    cognitive_state: Dict[str, Any] = {
        "action_history": [{"success": True} for _ in range(45)] + [{"success": False} for _ in range(15)],
        "strategy_stats": {
            "retrieve_context": {"successes": 90, "attempts": 120, "total_value": 40.0},
            "answer_with_tool": {"successes": 60, "attempts": 95, "total_value": 25.0},
            "structured_reasoning": {"successes": 50, "attempts": 70, "total_value": 22.0},
        },
        "action_log": [{"fallback_reason": None} for _ in range(40)]
        + [{"fallback_reason": "tool_error"} for _ in range(40)]
        + [{"fallback_reason": "tool_timeout"} for _ in range(20)],
        "tool_error_counts": {"answer_with_tool": 6},
        "last_sleep_phase_at": "2026-01-01T01:23:45+00:00",
        "pipeline_activity": [
            {
                "timestamp": "2026-01-01T01:30:00+00:00",
                "type": "control_update",
                "details": {"target": "low_confidence_threshold", "proposed_value": 0.35},
            }
        ],
    }
    return DecisionPipeline(cognitive_state)


def demo_dashboard() -> int:
    """Launch a live mini-demo with scripted proposals/actions and report export."""
    app = QApplication(sys.argv)
    pipeline = build_demo_pipeline()
    widget = ReflectionWidget(pipeline)
    widget.show()

    proposals: List[Dict[str, Any]] = [
        {
            "proposal_id": "p1",
            "change_type": "control_update",
            "target": "low_confidence_threshold",
            "proposed_value": 0.4,
            "confidence": 0.95,
        },
        {
            "proposal_id": "p2",
            "change_type": "tool_demotion",
            "target": "answer_with_tool",
            "confidence": 0.8,
        },
        {
            "proposal_id": "p3",
            "change_type": "unknown_type",
            "target": "dummy",
            "proposed_value": None,
            "confidence": 0.6,
        },
    ]

    widget.current_proposals = proposals
    widget._render_proposals()

    widget.action_handler.apply_selected([proposals[0]])
    widget.action_handler.reject_selected([proposals[2]])
    widget._refresh_dashboard()
    widget.export_report_to("ReflectionWidget_demo_report.xlsx")

    exec_fn = getattr(app, "exec", None) or getattr(app, "exec_", None)
    if exec_fn is None:
        raise RuntimeError("Unable to start Qt event loop")
    return int(exec_fn())


def main() -> int:
    """Launch Reflection Dashboard in standalone mode."""
    app = QApplication(sys.argv)
    widget = ReflectionWidget(build_demo_pipeline())
    widget.show()
    exec_fn = getattr(app, "exec", None) or getattr(app, "exec_", None)
    if exec_fn is None:
        raise RuntimeError("Unable to start Qt event loop: no exec/exec_ method found")
    return int(exec_fn())


if __name__ == "__main__":
    raise SystemExit(main())
